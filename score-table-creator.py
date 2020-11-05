from os import path
from numpy.core.fromnumeric import (
    shape,
    size
)
from openpyxl import (
    Workbook,
    load_workbook
)
from sys import exit
from time import sleep
from openpyxl.styles import numbers
from pandas import DataFrame
from openpyxl.utils.dataframe import dataframe_to_rows
from numpy import zeros
from pandas.core.arrays.sparse import dtype
from math import pow
from openpyxl.styles import (
    Font,
    Alignment,
    colors
)
from openpyxl.worksheet.table import Table
from openpyxl.utils import FORMULAE
from pandas import read_excel


def score_table():
    """
    The program is creating an excel sheet to keep player's scores for turn-based games that scores are needed to jot down.

    Has 3 choices to use;
    - Create a score table
    - Check the score table
    - Update the score table
    and 
    - Exit
    """

    msg = f"\n" + \
        ' Welcome to the "Score Table Creator" '.center(100, "*") + "\n"
    print(msg)
    game = str(
        input("What's the name of the game?: "))
    n_players = int(input("\nHow many players you'll play?: "))
    names = []
    for n in range(n_players):
        name = str(input(f"Please enter a name for player {n + 1}: ")).capitalize()
        names.append(name)
    filename = f"~\\Desktop\\{game}.xlsx"

    while True:
        choice = int(input(f"\n\
*-----------------------------*\n\
Create a new excel sheet -> 1\n\
Check the score table -> 2\n\
Update scores for the turn -> 3\n\
Exit -> 4\n\
*-----------------------------*\n\
What's your choice?: "))

        # ! creating a new excel sheet
        if choice == 1:
            wb = Workbook()
            sheet = wb.active
            sheet.title = game.upper()

            # ? creating table with given names
            np_zeros = zeros(
                shape=(int(pow(float(n_players), 2)), len(names)), dtype=int)
            data_dict = {"Players": names}
            key = [f'Turn {n}' for n in range(
                1, int(pow(float(n_players), 2)) + 1)]
            value = [np_zeros[n] for n in range(int(pow(float(n_players), 2)))]
            k_v = tuple(zip(key, value))
            data_dict.update(k_v)
            df = DataFrame(data_dict)

            # ? turning created table into dataframe for adding it to excel
            for row in dataframe_to_rows(df, index=False, header=True):
                sheet.append(row)

            # ? styling some cells
            sheet["A1"].font = Font(bold=True, size=12, italic=True)

            for row in range(2, n_players + 2):
                sheet[f"A{row}"].font = Font(
                    color="FF0000", size=11, bold=True, italic=True)

            for col in range(2, int(pow(float(n_players), 2)) + 2):
                sheet.cell(row=1, column=col).font = Font(
                    color="0000FF", size=12, bold=True)
                sheet.cell(row=1, column=col).alignment = Alignment(
                    horizontal="center")

            # ? creating total score column and styling it
            sheet.insert_cols(idx=1)
            sheet["A1"].value = "Total"
            sheet["A1"].font = Font(bold=True, size=13, color="00FF00")

            for row in range(2, n_players + 2):
                sheet[f"A{row}"].value = f"=SUM(C{row}:ZZ{row})"

            # ? create table filter
            tab = Table(displayName="players", ref=f"B1:B{n_players + 1}")
            sheet.add_table(tab)

            wb.save(path.expanduser(filename))

            print(f"\nThe score board for the game that is called \
{repr(game.capitalize())} is being created with {n_players} players, right now.")
            sleep(2.5)
            print(f"\nSetup has been done.\n\n\
Now, you have the excel file for {repr(game)} in your desktop.")
            sleep(2)

        # ! checking the score table
        elif choice == 2:
            read_df = read_excel(path.expanduser(
                filename), sheet_name=game.upper())
            print(read_df)

        # ! updating scores for a defined turn
        elif choice == 3:
            wb = load_workbook(path.expanduser(filename))
            sheet = wb.active

            turn = int(input("Which turn you're in?: ")) + 2
            for n in range(2, n_players + 2):
                sheet.cell(row=n, column=turn).value = int(
                    input(f"What is the score of {names[n - 2]} for turn {turn - 2}: "))

            wb.save(path.expanduser(filename))

            print(
                f"The score table of {game.capitalize()} is edited for turn {turn - 2}.")

        # ! exiting from program
        elif choice == 4:
            print("The program is closing...")
            sleep(1)
            exit()

        else:
            raise TypeError("Please type numbers, not letters")


if __name__ == "__main__":
    score_table()
