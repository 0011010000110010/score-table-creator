from os import path
from numpy.core.fromnumeric import shape
from openpyxl import Workbook, load_workbook
from sys import exit
from time import sleep
from pandas import DataFrame
from openpyxl.utils.dataframe import dataframe_to_rows
from numpy import zeros
from pandas.core.arrays.sparse import dtype


def game_points():

    msg = f"\n" + "Game Point Tracker".center(30, "*")
    game = str(input("Enter the name of the game: "))
    n_players = int(input("Enter player number: "))
    filename = f"~\\Desktop\\{game}.xlsx"

    while True:
        print(msg)
        choice = int(input(f"\n\
Create a new excel sheet -> 1\n\
Check the point table -> 2\n\
Add player names -> 3\n\
Add points for the turn -> 4\n\
Exit -> 5\n\n\
What's your choice?: "))

        if choice == 1:
            wb = Workbook()
            sheet = wb.active
            
            names = []
            for n in range(n_players):
                name = str(input("Enter a name: "))
                names.append(name)

            # turn = ""
            # n_turn = int(input("predicted turn number: "))
            # for t in range(n_turn):
            #     turn.append(f"Turn {t}")

            data = {
            "Players": names,
            "Turn 1": zeros(shape=(n_players), dtype=int),
            "Turn 2": zeros(shape=(n_players), dtype=int),
            "Turn 3": zeros(shape=(n_players), dtype=int),
            }
            
            df = DataFrame(data)
            for row in dataframe_to_rows(df, index=False, header=True):
                sheet.append(row)
            # sheet["A1"] = f"Welcome to the {repr(game)} point tracker"
            
            # for n in range(n_players):
                # sheet.cell(row=2, column=n).value = str(input("name: "))
            wb.save(path.expanduser(filename))

        elif choice == 2:
            loading_wb = load_workbook(path.expanduser(filename))
            sheet = loading_wb.active
            for row in sheet.rows:
                print(row.values)

        elif choice == 3:
            n_players = int(input("Enter player number: "))

            names = []
            for p in range(n_players):
                name = str(input("Enter a name: "))
                names.append(name)
                print(f"Player names are: {names}")
            break

        elif choice == 4:
            workbook = load_workbook(path.expanduser(filename))
            sheet = workbook.active
            round = int(input("Enter round number: "))
            for n in range(n_players):
                sheet.cell(row=round, column=n).value = input("point: ")
                workbook.save(path.expanduser(filename))
            break

        elif choice == 5:
            print("The program is closing...")
            sleep(4)
            exit()

        else:
            raise TypeError("Please type numbers, not letters")


if __name__ == "__main__":
    game_points()

"""
dict_list = []
names = ["ali", "ayşe", "veli", "canan"]
n_players = 4
zero = f"{zeros(shape=(n_players), dtype=int)}"

for key, value in dict_list.items():
    
for n in range(n_players):
    t = f"Turn {n}"
    turn.append(t)
    dict(turn, zero)

data = {"Players": names}

print(data)
"""